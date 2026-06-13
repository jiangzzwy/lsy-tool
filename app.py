"""Desktop GUI — 案件线索移送函批量生成工具

Single-panel step workflow: only current step is expanded,
previous steps collapse to a summary line, future steps are hidden.
"""

import os
import platform
import shutil
import subprocess
import sys
import threading
import traceback
from collections import Counter
from pathlib import Path

from dearpygui import dearpygui as dpg

import config
from api_client import BureauDB, MockApiClient
from excel_parser import parse_excel, split_rows
from word_generator import generate_word_docs
from ledger_generator import generate_ledger


# ── Colors ──────────────────────────────────────────────────────────────────

BG        = (22, 24, 30)
CARD_BG   = (30, 33, 42)
INPUT_BG  = (38, 41, 52)
INPUT_BD  = (52, 55, 68)
ACCENT    = (65, 135, 245)
ACCENT_H  = (85, 155, 255)
GREEN     = (75, 195, 125)
RED       = (230, 75, 75)
YELLOW    = (255, 200, 70)
TEXT      = (235, 237, 245)
TEXT2     = (155, 160, 178)
TEXT3     = (100, 105, 120)
DIM       = (60, 63, 76)
DIVIDER   = (44, 47, 58)
STEP_ON   = (65, 135, 245)
STEP_DONE = (55, 175, 115)
STEP_OFF  = (60, 63, 76)

W, H = 780, 640


# ── Helpers ──────────────────────────────────────────────────────────────────

def open_dir(path: str):
    p = Path(path)
    if not p.exists():
        return
    s = platform.system()
    try:
        if s == "Darwin":
            subprocess.Popen(["open", str(p)])
        elif s == "Windows":
            os.startfile(str(p))
        else:
            subprocess.Popen(["xdg-open", str(p)])
    except Exception:
        pass

def base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS)
    return Path(__file__).parent

def _log(msg: str):
    try:
        v = dpg.get_value("log_output")
        dpg.set_value("log_output", v + msg + "\n")
        dpg.set_y_scroll("log_wrap", -1.0)
    except Exception:
        pass

def _set_progress(pct: float, label: str = ""):
    try:
        dpg.set_value("progress", pct)
        dpg.set_value("progress_lbl", label)
    except Exception:
        pass

def _set_status(text: str, color=GREEN):
    try:
        dpg.set_value("status", text)
        dpg.configure_item("status", color=color)
    except Exception:
        pass

def _resolve_templates(src: str) -> dict:
    sp = Path(src)
    cands = [sp.parent, base_dir() / "demands"]
    out = {}
    for cat, rel in config.TEMPLATES.items():
        name = Path(rel).name
        for c in cands:
            t = c / name
            if t.exists():
                out[cat] = str(t)
                break
    return out

def _find_default_excel() -> str:
    d = base_dir() / "demands"
    if d.exists():
        for f in sorted(d.iterdir(), key=lambda x: x.name):
            if f.suffix in (".xlsx", ".xls") and "工单" in f.name and not f.name.startswith("~$"):
                return str(f)
    return ""

def _short_path(path_str: str, max_len: int = 50) -> str:
    """Show just filename or shortened path."""
    if not path_str:
        return ""
    p = Path(path_str)
    name = p.name
    if len(name) <= max_len:
        return name
    return "..." + name[-(max_len - 3):]

def _show_step(n: int):
    """Switch visible panel to step n, hide others."""
    st.current_step = n
    for i in range(1, 4):
        dpg.configure_item(f"step_panel_{i}", show=(i == n))
        # Step dots
        color = STEP_DONE if i < n else (STEP_ON if i == n else STEP_OFF)
        dpg.configure_item(f"step_circle_{i}", color=color)
        dpg.configure_item(f"step_label_{i}", color=color if i <= n else STEP_OFF)
    # Update summary lines for completed steps
    if n >= 2:
        src = dpg.get_value("source_excel")
        dpg.set_value("step1_summary", f"预检完成 | {_short_path(src)}")
        dpg.configure_item("step1_summary", show=True)
    if n >= 3:
        dpg.set_value("step2_summary", "数据补充完成")
        dpg.configure_item("step2_summary", show=True)


# ── State ────────────────────────────────────────────────────────────────────

class State:
    running = False
    parsed_items = None
    parsed_rows = None
    db_count = 0
    current_step = 1

st = State()


# ── Worker: Pre-check ────────────────────────────────────────────────────────

def run_precheck():
    try:
        src = dpg.get_value("source_excel")
        out = dpg.get_value("output_dir")
        if not src or not Path(src).exists():
            _log("ERROR: 源文件未找到")
            _set_status("文件未找到", RED)
            return
        if not out:
            out = str(base_dir() / "output")
        Path(out).mkdir(parents=True, exist_ok=True)

        _set_status("预检中...", YELLOW)
        _set_progress(0.15, "解析 Excel...")
        _log("解析 Excel...")
        rows = parse_excel(src)
        stats = Counter(r.classification for r in rows)
        for k, v in stats.items():
            _log(f"  {k}: {v}")

        _set_progress(0.35, "拆分行...")
        _log("拆分行...")
        items = split_rows(rows)
        _log(f"  共 {len(items)} 项")

        st.parsed_items = items
        st.parsed_rows = rows

        _set_progress(0.55, "检查数据库...")
        _log("检查本地数据库...")
        db = BureauDB()
        st.db_count = db.count()
        coverage = db.check_coverage(items)
        db.close()

        _log(f"  数据库: {st.db_count} 条记录")
        _log(f"  需要登记机关: {coverage['total']} 项")
        _log(f"  已覆盖: {coverage['covered']} 项 ({coverage['coverage_pct']:.0f}%)")
        _log(f"  缺失: {len(coverage['missing'])} 项")

        dpg.set_value("s_db", str(st.db_count))
        dpg.set_value("s_need", str(coverage["total"]))
        dpg.set_value("s_ok", str(coverage["covered"]))
        dpg.set_value("s_miss", str(len(coverage["missing"])))

        if coverage["missing"]:
            tpl_path = Path(out) / "待补充登记机关.xlsx"
            db2 = BureauDB()
            db2.export_template(coverage["missing"], str(tpl_path))
            db2.close()
            _log(f"\n  已生成补充模板: 待补充登记机关.xlsx")
            _log(f"  请在模板中填写登记机关列，保存后进入下一步导入")

            dpg.configure_item("s_miss", color=RED)
            dpg.set_value("step2_hint", f"缺失 {len(coverage['missing'])} 项登记机关，请导入补充数据")
            _set_progress(0.6, f"缺失 {len(coverage['missing'])} 项")
            _set_status(f"缺失 {len(coverage['missing'])} 项", YELLOW)
            _show_step(2)
        else:
            _log("所有登记机关已覆盖!")
            dpg.configure_item("s_miss", color=GREEN)
            _set_progress(1.0, "就绪")
            _set_status("预检通过", GREEN)
            _show_step(3)

    except Exception as e:
        _log(f"FATAL: {e}")
        _log(traceback.format_exc())
        _set_status("预检失败", RED)
        _show_step(1)
    finally:
        st.running = False
        dpg.configure_item("btn_precheck", enabled=True)


# ── Worker: Import ───────────────────────────────────────────────────────────

def run_import():
    try:
        import_path = dpg.get_value("import_path")
        if not import_path or not Path(import_path).exists():
            _log("ERROR: 补充数据文件未找到")
            _set_status("文件未找到", RED)
            return

        _set_status("导入中...", YELLOW)
        _set_progress(0.3, "导入数据...")

        db = BureauDB()
        result = db.import_template(import_path)
        db.close()

        _log(f"导入完成: {result['imported']} 条新增, {result['skipped']} 条跳过")
        if result["errors"]:
            for e in result["errors"]:
                _log(f"  错误: {e}")

        st.db_count = BureauDB().count()
        dpg.set_value("s_db", str(st.db_count))
        _set_progress(0.6, "重新检查...")

        if st.parsed_items:
            db2 = BureauDB()
            coverage = db2.check_coverage(st.parsed_items)
            db2.close()

            _log(f"重新检查: {coverage['covered']}/{coverage['total']} 已覆盖")
            _log(f"剩余缺失: {len(coverage['missing'])}")

            dpg.set_value("s_ok", str(coverage["covered"]))
            dpg.set_value("s_miss", str(len(coverage["missing"])))

            if coverage["missing"]:
                _set_status(f"仍有 {len(coverage['missing'])} 项缺失", YELLOW)
                dpg.configure_item("s_miss", color=RED)
                out = dpg.get_value("output_dir") or str(base_dir() / "output")
                tpl_path = Path(out) / "待补充登记机关.xlsx"
                db3 = BureauDB()
                db3.export_template(coverage["missing"], str(tpl_path))
                db3.close()
                _log(f"已重新生成模板: 待补充登记机关.xlsx")
                dpg.set_value("step2_hint", f"仍有 {len(coverage['missing'])} 项缺失，请继续补充")
                _show_step(2)
            else:
                _set_status("全部覆盖，可以生成", GREEN)
                dpg.configure_item("s_miss", color=GREEN)
                _set_progress(1.0, "就绪")
                _show_step(3)

    except Exception as e:
        _log(f"FATAL: {e}")
        _log(traceback.format_exc())
        _set_status("导入失败", RED)
    finally:
        st.running = False
        dpg.configure_item("btn_import", enabled=True)


# ── Worker: Generate ─────────────────────────────────────────────────────────

def run_generation():
    try:
        out = dpg.get_value("output_dir")
        if not out:
            out = str(base_dir() / "output")
        Path(out).mkdir(parents=True, exist_ok=True)

        if not st.parsed_items:
            _log("ERROR: 请先执行预检")
            _set_status("请先预检", RED)
            return

        items = st.parsed_items
        tpls = _resolve_templates(dpg.get_value("source_excel"))
        if not tpls:
            _log("ERROR: 模板未找到")
            return

        db = BureauDB()
        _set_status("生成中...", YELLOW)
        _set_progress(0.05, "生成 Word...")
        _log("生成 Word 文档...")

        total = len(items)
        word_results = []
        for i, item in enumerate(items):
            cls = item.row.classification
            template_path = tpls.get(cls)
            if not template_path:
                continue
            try:
                from word_generator import _fill_template, _build_doc_filename
                doc = _fill_template(template_path, item, db)
                filename = _build_doc_filename(item)
                cls_dir = Path(out) / cls
                cls_dir.mkdir(parents=True, exist_ok=True)
                doc.save(str(cls_dir / f"{filename}.docx"))
                word_results.append({"filename": "ok"})
            except Exception as e:
                word_results.append({"error": str(e)})

            if (i + 1) % 20 == 0 or i == total - 1:
                pct = 0.05 + 0.85 * (i + 1) / total
                _set_progress(pct, f"Word {i+1}/{total}")

        ok = sum(1 for r in word_results if r.get("filename"))
        err = sum(1 for r in word_results if r.get("error"))
        _log(f"  Word: {ok} 成功, {err} 失败")

        _set_progress(0.92, "生成台账...")
        _log("生成台账...")
        lp = generate_ledger(items, out, db)
        _log(f"  台账: {lp}")
        db.close()

        _set_progress(1.0, "完成")
        _set_status(f"完成: {ok} Word + 1 台账", GREEN)

        dpg.set_value("s_word", str(ok))
        dpg.set_value("s_ledger", "1")
        dpg.set_value("s_err", str(err))
        dpg.configure_item("s_err", color=RED if err > 0 else GREEN)

        cls_parts = "  ".join(
            f"{k}:{v}" for k, v in Counter(
                i.row.classification for i in items
            ).items() if k != "其他"
        )
        dpg.set_value("s_cls", cls_parts)
        dpg.configure_item("s_cls", color=TEXT2)

    except Exception as e:
        _log(f"FATAL: {e}")
        _log(traceback.format_exc())
        _set_status("生成失败", RED)
    finally:
        st.running = False
        dpg.configure_item("btn_gen", enabled=True)


# ── Callbacks ────────────────────────────────────────────────────────────────

def cb_precheck(s, d):
    if st.running:
        return
    st.running = True
    dpg.configure_item("btn_precheck", enabled=False)
    dpg.set_value("log_output", "")
    threading.Thread(target=run_precheck, daemon=True).start()

def cb_import(s, d):
    if st.running:
        return
    st.running = True
    dpg.configure_item("btn_import", enabled=False)
    threading.Thread(target=run_import, daemon=True).start()

def cb_gen(s, d):
    if st.running:
        return
    st.running = True
    dpg.configure_item("btn_gen", enabled=False)
    threading.Thread(target=run_generation, daemon=True).start()

def cb_open(s, d):
    o = dpg.get_value("output_dir") or str(base_dir() / "output")
    if Path(o).exists():
        open_dir(o)

def cb_clr(s, d):
    dpg.set_value("log_output", "")

def cb_bsrc(s, d):
    if d:
        dpg.set_value("source_excel", d["file_path_name"])

def cb_bout(s, d):
    if d:
        dpg.set_value("output_dir", d["file_path_name"])

def cb_bimp(s, d):
    if d:
        dpg.set_value("import_path", d["file_path_name"])

def cb_export_db(s, d):
    import openpyxl
    db = BureauDB()
    rows = db._db.execute(
        "SELECT credit_code, bureau, company, source, updated_at FROM bureau_cache ORDER BY credit_code"
    ).fetchall()
    db.close()
    if not rows:
        _log("数据库中没有记录")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业登记机关"
    headers = ["统一社会信用代码", "登记机关", "企业名称", "来源", "更新时间"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    for row_idx, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row_idx, col, val)
    col_widths = [22, 30, 25, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    out_dir = dpg.get_value("output_dir") or str(base_dir() / "output")
    dest = Path(out_dir) / "企业登记机关数据库.xlsx"
    wb.save(str(dest))
    wb.close()
    _log(f"已导出: {dest} ({len(rows)} 条)")
    open_dir(out_dir)

def cb_reset(s, d):
    st.parsed_items = None
    st.parsed_rows = None
    dpg.set_value("log_output", "")
    _set_progress(0.0, "")
    _set_status("就绪", GREEN)
    dpg.set_value("s_db", "0")
    dpg.set_value("s_need", "--")
    dpg.set_value("s_ok", "--")
    dpg.set_value("s_miss", "--")
    dpg.set_value("s_word", "--")
    dpg.set_value("s_ledger", "--")
    dpg.set_value("s_err", "--")
    dpg.set_value("s_cls", "--")
    dpg.configure_item("s_miss", color=YELLOW)
    dpg.configure_item("s_err", color=GREEN)
    dpg.configure_item("s_cls", color=TEXT3)
    dpg.configure_item("step1_summary", show=False)
    dpg.configure_item("step2_summary", show=False)
    _show_step(1)


# ── Build GUI ────────────────────────────────────────────────────────────────

def build_gui():
    dpg.create_context()

    # ── Fonts ──
    fps = {
        "Darwin": ["/System/Library/Fonts/PingFang.ttc"],
        "Windows": ["C:/Windows/Fonts/msyh.ttc"],
    }.get(platform.system(), [])

    with dpg.font_registry():
        f14 = f18 = f24 = None
        for fp in fps:
            if Path(fp).exists():
                f14 = dpg.add_font(fp, 13)
                f18 = dpg.add_font(fp, 16)
                f24 = dpg.add_font(fp, 22)
                break

    # ── File dialogs ──
    with dpg.file_dialog(directory_selector=False, show=False,
                          callback=cb_bsrc, tag="dlg_src", width=700, height=400):
        dpg.add_file_extension(".xlsx", label=".xlsx")
        dpg.add_file_extension(".xls", label=".xls")

    with dpg.file_dialog(directory_selector=True, show=False,
                          callback=cb_bout, tag="dlg_out", width=700, height=400):
        pass

    with dpg.file_dialog(directory_selector=False, show=False,
                          callback=cb_bimp, tag="dlg_imp", width=700, height=400):
        dpg.add_file_extension(".xlsx", label=".xlsx")

    # ── Theme ──
    with dpg.theme() as theme:
        with dpg.theme_component(dpg.mvAll):
            dpg.add_theme_color(dpg.mvThemeCol_WindowBg, BG, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_ChildBg, CARD_BG, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Border, DIVIDER, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Separator, DIVIDER, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, INPUT_BG, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_FrameBgHovered, INPUT_BD, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_FrameBgActive, INPUT_BD, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Button, ACCENT, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, ACCENT_H, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (50, 115, 220), category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Text, TEXT, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_TextDisabled, DIM, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_ScrollbarBg, BG, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_ScrollbarGrab, INPUT_BD, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_PopupBg, CARD_BG, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Header, INPUT_BG, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_HeaderHovered, INPUT_BD, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_HeaderActive, INPUT_BD, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_PlotHistogram, ACCENT, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_PlotHistogramHovered, ACCENT_H, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 4, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_WindowRounding, 0, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_ChildRounding, 8, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_PopupRounding, 4, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FramePadding, 8, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_ItemSpacing, 8, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_WindowPadding, 20, 14, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_ScrollbarSize, 8, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_WindowBorderSize, 0, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameBorderSize, 0, category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_GrabRounding, 3, category=dpg.mvThemeCat_Core)

    dpg.bind_theme(theme)

    # ── Main window ──
    with dpg.window(tag="main_win", no_title_bar=True, no_scrollbar=False):

        # ══ Step Indicator ══════════════════════════════════════════════
        with dpg.group(horizontal=True):
            dpg.add_spacer(width=10)
            # Step 1 circle
            with dpg.group(horizontal=True):
                dpg.add_text("1", tag="step_circle_1", color=STEP_ON)
                dpg.add_text("预检", tag="step_label_1", color=STEP_ON)
            dpg.add_spacer(width=16)
            dpg.add_text("---", color=DIM)
            dpg.add_spacer(width=16)
            # Step 2
            with dpg.group(horizontal=True):
                dpg.add_text("2", tag="step_circle_2", color=STEP_OFF)
                dpg.add_text("补充数据", tag="step_label_2", color=STEP_OFF)
            dpg.add_spacer(width=16)
            dpg.add_text("---", color=DIM)
            dpg.add_spacer(width=16)
            # Step 3
            with dpg.group(horizontal=True):
                dpg.add_text("3", tag="step_circle_3", color=STEP_OFF)
                dpg.add_text("生成文件", tag="step_label_3", color=STEP_OFF)
            # Status far right
            dpg.add_spacer(width=-1)
            dpg.add_text(tag="status", default_value="就绪", color=GREEN)

        dpg.add_separator()
        dpg.add_spacer(height=6)

        # ══ Step 1 Panel: Pre-check ════════════════════════════════════
        with dpg.child_window(tag="step_panel_1", height=130, border=True):
            dpg.add_text("选择源文件并预检", color=TEXT)
            dpg.add_spacer(height=4)

            with dpg.group(horizontal=True):
                dpg.add_text("源文件", color=TEXT2)
                dpg.add_input_text(tag="source_excel", width=-300,
                                   default_value=_find_default_excel())
                dpg.add_button(label="浏览", width=60,
                               callback=lambda: dpg.show_item("dlg_src"))

            with dpg.group(horizontal=True):
                dpg.add_text("输出到", color=TEXT2)
                dpg.add_input_text(tag="output_dir", width=-300,
                                   default_value=str(base_dir() / "output"))
                dpg.add_button(label="浏览", width=60,
                               callback=lambda: dpg.show_item("dlg_out"))

            dpg.add_spacer(height=4)
            dpg.add_button(label="  开始预检  ", tag="btn_precheck",
                           callback=cb_precheck, height=34, width=180)

        # ══ Step 1 Summary (shown after completion) ════════════════════
        dpg.add_text(tag="step1_summary", default_value="", color=STEP_DONE, show=False)

        # ══ Step 2 Panel: Import ═══════════════════════════════════════
        with dpg.child_window(tag="step_panel_2", height=130, border=True, show=False):
            dpg.add_text(tag="step2_hint", default_value="缺失登记机关，请导入补充数据", color=YELLOW)
            dpg.add_spacer(height=4)

            with dpg.group(horizontal=True):
                dpg.add_text("补充文件", color=TEXT2)
                dpg.add_input_text(tag="import_path", width=-300,
                                   default_value="", hint="选择已填写的补充模板...")
                dpg.add_button(label="浏览", width=60,
                               callback=lambda: dpg.show_item("dlg_imp"))

            dpg.add_spacer(height=4)
            with dpg.group(horizontal=True):
                dpg.add_button(label="  导入数据  ", tag="btn_import",
                               callback=cb_import, height=34, width=180)
                dpg.add_spacer(width=12)
                dpg.add_button(label=" 打开输出目录 ", callback=cb_open,
                               height=34, width=140)

        # ══ Step 2 Summary ═════════════════════════════════════════════
        dpg.add_text(tag="step2_summary", default_value="", color=STEP_DONE, show=False)

        # ══ Step 3 Panel: Generate ═════════════════════════════════════
        with dpg.child_window(tag="step_panel_3", height=100, border=True, show=False):
            dpg.add_text("所有登记机关已覆盖，可以生成文件", color=GREEN)
            dpg.add_spacer(height=6)

            with dpg.group(horizontal=True):
                dpg.add_button(label="  开始生成  ", tag="btn_gen",
                               callback=cb_gen, height=34, width=180)
                dpg.add_spacer(width=16)
                dpg.add_progress_bar(tag="progress", default_value=0.0,
                                     width=-180, height=18)
                dpg.add_text(tag="progress_lbl", default_value="", color=TEXT2)

        dpg.add_spacer(height=8)

        # ══ Stats Cards ════════════════════════════════════════════════
        with dpg.group(horizontal=True):
            # Card: DB
            with dpg.child_window(width=130, height=52, border=True):
                dpg.add_text("数据库", color=TEXT3)
                dpg.add_text(tag="s_db", default_value="0", color=ACCENT)
            dpg.add_spacer(width=6)
            # Card: Need
            with dpg.child_window(width=100, height=52, border=True):
                dpg.add_text("需要", color=TEXT3)
                dpg.add_text(tag="s_need", default_value="--", color=TEXT2)
            dpg.add_spacer(width=6)
            # Card: Covered
            with dpg.child_window(width=100, height=52, border=True):
                dpg.add_text("已覆盖", color=TEXT3)
                dpg.add_text(tag="s_ok", default_value="--", color=GREEN)
            dpg.add_spacer(width=6)
            # Card: Missing
            with dpg.child_window(width=100, height=52, border=True):
                dpg.add_text("缺失", color=TEXT3)
                dpg.add_text(tag="s_miss", default_value="--", color=YELLOW)
            dpg.add_spacer(width=6)
            # Card: Word
            with dpg.child_window(width=80, height=52, border=True):
                dpg.add_text("Word", color=TEXT3)
                dpg.add_text(tag="s_word", default_value="--", color=GREEN)
            dpg.add_spacer(width=6)
            # Card: Ledger
            with dpg.child_window(width=70, height=52, border=True):
                dpg.add_text("台账", color=TEXT3)
                dpg.add_text(tag="s_ledger", default_value="--", color=GREEN)
            dpg.add_spacer(width=6)
            # Card: Error
            with dpg.child_window(width=70, height=52, border=True):
                dpg.add_text("错误", color=TEXT3)
                dpg.add_text(tag="s_err", default_value="--", color=GREEN)

        dpg.add_spacer(height=6)

        # ══ Log Section ════════════════════════════════════════════════
        with dpg.group(horizontal=True):
            dpg.add_text("日志", color=TEXT2)
            dpg.add_spacer(width=-1)
            dpg.add_button(label=" 导出数据 ", callback=cb_export_db, height=24, width=85)
            dpg.add_button(label=" 重置 ", callback=cb_reset, height=24, width=55)
            dpg.add_button(label=" 清空 ", callback=cb_clr, height=24, width=55)

        with dpg.child_window(tag="log_wrap", border=False, height=-1):
            dpg.add_input_text(tag="log_output", multiline=True,
                               width=-1, height=-1,
                               readonly=True, default_value="")

    # ── Font ──
    if f14:
        dpg.bind_font(f14)

    # ── Viewport ──
    dpg.create_viewport(title="案件线索移送函生成工具", width=W, height=H)
    dpg.setup_dearpygui()
    dpg.set_viewport_clear_color(BG)
    dpg.show_viewport()
    dpg.set_primary_window("main_win", True)
    dpg.start_dearpygui()
    dpg.destroy_context()


if __name__ == "__main__":
    build_gui()
