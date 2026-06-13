"""Parse source Excel and classify each row."""

import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class RowData:
    """Parsed data for a single Excel row (possibly split into multiple docs)."""
    row_num: int
    task_no: str           # B列 任务单号
    reg_no: str            # C列 登记单号
    name: str              # F列 姓名
    phone: str             # G列 手机号码
    biz_type: str          # H列 三方/自营
    order_nos: list[str]   # I列 订单号 (split by comma)
    shop_names: list[str]  # L列 店铺名称 (split by comma)
    category: str          # M列 商品类别
    company: str           # N列 企业名称
    address: str           # O列 企业地址
    credit_code: str       # P列 统一社会信用代码
    # Derived
    classification: str = ""  # 外卖/三方公司/三方个人/自营开票/其他

    @property
    def task_no_str(self) -> str:
        return str(self.task_no).strip()

    @property
    def task_no_prefix4(self) -> str:
        """First 4 digits of task number (任务单号前4位)."""
        s = self.task_no_str
        return s[:4] if len(s) >= 4 else s


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _split_comma(val: str) -> list[str]:
    """Split by comma (both Chinese and English), return non-empty items."""
    items = val.replace("，", ",").split(",")
    return [i.strip() for i in items if i.strip()]


def classify_row(row: RowData) -> str:
    """Classify a row into one of: 外卖, 三方公司, 三方个人, 自营开票, 其他."""
    h = row.biz_type
    m = row.category
    l_str = ",".join(row.shop_names)
    p = row.credit_code

    if m.startswith("外卖美食") or m.startswith("外卖"):
        return "外卖"

    if h in ("三方", "三方/自营"):
        if p and p not in ("暂无", "/", "None", ""):
            return "三方公司"
        else:
            return "三方个人"

    if h == "自营":
        return "自营开票"

    if h == "三方/自营" and "自营" in l_str:
        return "自营开票"

    if not h or h == "/":
        return "其他"

    return "其他"


def parse_excel(excel_path: str | Path) -> list[RowData]:
    """Parse source Excel and return list of RowData."""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active
    rows = []

    for row_num in range(2, ws.max_row + 1):
        task_no = _safe_str(ws.cell(row_num, 2).value)
        if not task_no or task_no == "/":
            continue

        row = RowData(
            row_num=row_num,
            task_no=task_no,
            reg_no=_safe_str(ws.cell(row_num, 3).value),
            name=_safe_str(ws.cell(row_num, 6).value),
            phone=_safe_str(ws.cell(row_num, 7).value),
            biz_type=_safe_str(ws.cell(row_num, 8).value),
            order_nos=_split_comma(_safe_str(ws.cell(row_num, 9).value)),
            shop_names=_split_comma(_safe_str(ws.cell(row_num, 12).value)),
            category=_safe_str(ws.cell(row_num, 13).value),
            company=_safe_str(ws.cell(row_num, 14).value),
            address=_safe_str(ws.cell(row_num, 15).value),
            credit_code=_safe_str(ws.cell(row_num, 16).value),
        )
        row.classification = classify_row(row)
        rows.append(row)

    wb.close()
    logger.info(f"Parsed {len(rows)} data rows from {excel_path}")
    return rows


@dataclass
class SplitItem:
    """One split unit from a RowData: one shop + corresponding orders."""
    row: RowData
    shop_name: str         # single shop name
    order_nos: list[str]   # orders for this shop
    index: int = 0         # split index (1-based), 0 means no suffix needed
    total: int = 1         # total splits for this row


def split_rows(rows: list[RowData]) -> list[SplitItem]:
    """Apply splitting rules: when multiple shops, split into one doc per shop."""
    items = []
    for row in rows:
        if row.classification == "其他":
            continue

        shops = row.shop_names if row.shop_names else [""]
        orders = row.order_nos if row.order_nos else []

        if len(shops) == 1:
            # Single shop: no split, one doc with all orders
            items.append(SplitItem(
                row=row,
                shop_name=shops[0],
                order_nos=orders,
                index=0,
                total=1,
            ))
        else:
            # Multiple shops: split, one doc per shop
            for i, shop in enumerate(shops):
                items.append(SplitItem(
                    row=row,
                    shop_name=shop,
                    order_nos=orders,
                    index=i + 1,
                    total=len(shops),
                ))
    return items
