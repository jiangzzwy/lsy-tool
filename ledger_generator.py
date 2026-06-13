"""Generate the ledger Excel (台账) from parsed data."""

import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from api_client import BureauDB
from excel_parser import RowData, SplitItem, split_rows

logger = logging.getLogger(__name__)


def _today_ledger_str() -> str:
    """Return today's date for ledger: e.g. 2026.6.10"""
    d = date.today()
    return f"{d.year}.{d.month}.{d.day}"


def _current_year() -> str:
    return str(date.today().year)


def _format_orders(orders: list[str]) -> str:
    """Join order numbers with Chinese comma separator."""
    return "、".join(orders)


def _build_sales_desc(item: SplitItem) -> str:
    """Build column E (销售商品) content based on classification."""
    row = item.row
    cls = row.classification
    orders_str = _format_orders(item.order_nos)
    shop = item.shop_name

    if cls == "外卖":
        return f"销售订单编号{orders_str}商品"

    if cls == "三方公司":
        return f'第三方商家"{shop}"销售订单编号{orders_str}商品'

    if cls == "三方个人":
        return f'第三方商家"{shop}"销售订单编号{orders_str}商品'

    if cls == "自营开票":
        return f"销售订单编号{orders_str}商品"

    return ""


def _get_bureau(item: SplitItem, db: BureauDB) -> str:
    """Get bureau name - reuse logic from word_generator."""
    from word_generator import _get_bureau as wg_get_bureau
    return wg_get_bureau(item, db)


def generate_ledger(
    items: list[SplitItem],
    output_dir: str | Path,
    db: BureauDB,
) -> str:
    """Generate ledger Excel file. Returns output file path."""
    output_path = Path(output_dir)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "台账"

    # Headers (A-P) — matching template layout exactly
    headers = [
        "电话",          # A
        "姓名",          # B
        "",              # C (冒号)
        "",              # D (事项固定内容)
        '"XXXX"在销售订单编号（或商品编号）xxxx商品时，',  # E
        "",              # F (涉嫌违法...涉案主体为)
        "公司（主体）名称", # G
        "",              # H (法规固定内容)
        "XX市场监督管理局", # I
        "",              # J (移交函号前缀)
        "任务单号",       # K
        "",              # L (号）)
        "",              # M (结果固定内容)
        "落款日期",       # N
        "备注",           # O
        "特殊备注",       # P
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        if header:
            ws.cell(1, col, header)

    # Fixed content columns — matching template exactly
    COL_C = "："                                           # C - 冒号
    COL_D = "我单位收到您对京东商城"                        # D - 事项
    COL_F = "涉嫌违法等问题的举报。经调查，上述商品涉案主体为"  # F - 合并内容
    COL_H = "，根据《市场监督管理行政处罚程序规定》第十条平台内经营者的违法行为由其实际经营地县级以上市场监督管理部门管辖，特将该举报移送至"  # H - 法规
    COL_J = f"处理（移交函号：京技管商务消移送〔{_current_year()}〕J"  # J prefix
    COL_L = "号）"                                         # L
    COL_M = "。现有证据无法认定被举报方违法行为成立，依据《市场监督管理行政处罚程序规定》第二十条第一款第（四）项的规定，我单位决定不予立案。特此告知！【北京经济技术开发区管理委员会】"  # M

    # Write data rows
    row_idx = 2
    for item in items:
        row = item.row
        cls = row.classification

        # Split rows: subsequent splits have empty phone/name
        is_subsequent_split = item.index > 1

        ws.cell(row_idx, 1, "" if is_subsequent_split else row.phone)         # A - 手机号码
        ws.cell(row_idx, 2, "" if is_subsequent_split else row.name.replace("\n", ""))  # B - 姓名
        ws.cell(row_idx, 3, COL_C)                                            # C - 冒号
        ws.cell(row_idx, 4, COL_D)                                            # D - 事项
        ws.cell(row_idx, 5, _build_sales_desc(item))                          # E - 销售商品
        ws.cell(row_idx, 6, COL_F)                                            # F - 涉嫌违法...涉案主体为
        ws.cell(row_idx, 7, row.company)                                      # G - 公司名称
        ws.cell(row_idx, 8, COL_H)                                            # H - 法规
        ws.cell(row_idx, 9, _get_bureau(item, db))                            # I - 市监管

        # J - 移交函号前缀 (only on first split or non-split rows)
        if item.index <= 1:
            ws.cell(row_idx, 10, COL_J)
        # K - 任务单号 (with split suffix)
        task_no_for_ledger = row.task_no_str
        if item.index > 0:
            task_no_for_ledger = f"{row.task_no_str}-{item.index}"
        ws.cell(row_idx, 11, task_no_for_ledger)
        ws.cell(row_idx, 12, COL_L)                                           # L - 号）
        ws.cell(row_idx, 13, COL_M)                                           # M - 结果
        ws.cell(row_idx, 14, _today_ledger_str())                             # N - 落款日期
        ws.cell(row_idx, 15, None)                                            # O - 备注
        ws.cell(row_idx, 16, None)                                            # P - 特殊备注

        row_idx += 1

    # Adjust column widths
    col_widths = {
        1: 15,   # A 电话
        2: 12,   # B 姓名
        3: 3,    # C 冒号
        4: 20,   # D 事项
        5: 40,   # E 销售商品
        6: 35,   # F 涉嫌违法...涉案主体为
        7: 25,   # G 公司名称
        8: 55,   # H 法规
        9: 30,   # I 市监管
        10: 35,  # J 移交函号前缀
        11: 18,  # K 任务单号
        12: 5,   # L 号）
        13: 55,  # M 结果
        14: 12,  # N 落款日期
        15: 10,  # O 备注
        16: 15,  # P 特殊备注
    }
    for col, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    out_file = output_path / "台账.xlsx"
    wb.save(str(out_file))
    wb.close()
    logger.info(f"Generated ledger: {out_file} ({row_idx - 2} rows)")
    return str(out_file)
