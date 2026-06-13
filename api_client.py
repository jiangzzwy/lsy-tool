"""Enterprise registration authority (登记机关) lookup via local SQLite database.

No API calls. Users maintain a local bureau database by:
1. Pre-checking source Excel against the DB
2. Exporting missing credit codes as a template Excel
3. Filling in the template and importing it back
4. Once all codes are covered, generation can proceed

Address heuristic (mock) is used as a fallback for rows without credit codes.
"""

import logging
import re
import sqlite3
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

DB_FILE = Path(__file__).parent / "output" / ".bureau_cache.db"


def get_db() -> sqlite3.Connection:
    """Get or create the SQLite bureau database."""
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_FILE))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS bureau_cache (
            credit_code TEXT PRIMARY KEY,
            bureau TEXT NOT NULL,
            company TEXT DEFAULT '',
            source TEXT NOT NULL DEFAULT 'import',
            updated_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
    """)
    conn.commit()
    return conn


class BureauDB:
    """Manage the local bureau database: check, export, import."""

    def __init__(self):
        self._db = get_db()

    def close(self):
        self._db.close()

    # ── Query ──────────────────────────────────────────────────────────────

    def lookup(self, credit_code: str) -> str | None:
        """Look up bureau by credit code. Returns bureau name or None."""
        if not credit_code or credit_code in ("暂无", "/", "None", ""):
            return None
        cc = credit_code.strip()
        try:
            row = self._db.execute(
                "SELECT bureau FROM bureau_cache WHERE credit_code = ?",
                (cc,),
            ).fetchone()
            return row[0] if row else None
        except Exception:
            return None

    def count(self) -> int:
        """Total entries in DB."""
        row = self._db.execute("SELECT COUNT(*) FROM bureau_cache").fetchone()
        return row[0] if row else 0

    # ── Pre-check ──────────────────────────────────────────────────────────

    def check_coverage(self, items: list) -> dict:
        """Check which SplitItems have bureau coverage.

        Returns dict with:
            total: total items needing bureau (三方公司 + 自营开票)
            covered: items with bureau in DB
            missing: list of (credit_code, company, classification) tuples
        """
        missing = []
        total = 0
        covered = 0

        seen_codes = set()  # avoid duplicates

        for item in items:
            cls = item.row.classification
            if cls not in ("三方公司", "自营开票"):
                continue

            total += 1
            cc = (item.row.credit_code or "").strip()
            if cc in ("", "暂无", "/", "None"):
                # No credit code — will use address heuristic
                continue

            if cc in seen_codes:
                covered += 1
                continue
            seen_codes.add(cc)

            bureau = self.lookup(cc)
            if bureau:
                covered += 1
            else:
                missing.append((cc, item.row.company or "", cls))

        return {
            "total": total,
            "covered": covered,
            "missing": missing,
            "coverage_pct": (covered / total * 100) if total > 0 else 100,
        }

    # ── Export template ────────────────────────────────────────────────────

    def export_template(self, missing: list, output_path: str) -> str:
        """Export missing credit codes as a fillable Excel template.

        missing: list of (credit_code, company, classification) tuples
        Returns the output file path.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "登记机关补充"

        # Headers
        headers = ["统一社会信用代码", "企业名称", "分类", "登记机关（请填写）"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for i, (cc, company, cls) in enumerate(missing, 2):
            ws.cell(i, 1, cc)
            ws.cell(i, 2, company)
            ws.cell(i, 3, cls)
            ws.cell(i, 4, "")  # User fills this

        # Column widths
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 35

        wb.save(output_path)
        return output_path

    # ── Import template ────────────────────────────────────────────────────

    def import_template(self, template_path: str) -> dict:
        """Import a filled template Excel into the database.

        Returns dict with:
            imported: number of new entries
            skipped: number of entries with empty bureau
            errors: list of error messages
        """
        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb.active

        imported = 0
        skipped = 0
        errors = []

        for row_num in range(2, ws.max_row + 1):
            cc = str(ws.cell(row_num, 1).value or "").strip()
            company = str(ws.cell(row_num, 2).value or "").strip()
            bureau = str(ws.cell(row_num, 4).value or "").strip()

            if not cc:
                continue

            if not bureau:
                skipped += 1
                continue

            try:
                self._db.execute(
                    """INSERT OR REPLACE INTO bureau_cache (credit_code, bureau, company, source, updated_at)
                       VALUES (?, ?, ?, 'import', datetime('now', 'localtime'))""",
                    (cc, bureau, company),
                )
                self._db.commit()
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {e}")

        wb.close()
        return {"imported": imported, "skipped": skipped, "errors": errors}


class MockApiClient:
    """Derive bureau from address using heuristic regex matching.
    Used as fallback for rows without credit codes (三方个人, or rows with 暂无)."""

    SPECIAL_RULES = {
        "北京经济技术开发区": "北京市北京经济技术开发区市场监督管理局",
        "吉林高新技术产业开发区": "吉林市市场监督管理局吉林高新技术产业开发区分局",
        "吉林市高新区": "吉林市市场监督管理局吉林高新技术产业开发区分局",
        "上海京东到家友恒": "上海市杨浦区市场监督管理局",
        "潮安县": "潮州市潮安区市场监督管理局",
        "潮安区": "潮州市潮安区市场监督管理局",
        "嘉定工业区": "上海市嘉定区市场监督管理局",
        "前海深港合作区": "深圳市市场监督管理局南山监管局",
        "东湖新技术开发区": "武汉市东湖新技术开发区市场监督管理局",
        "望城经济技术开发区": "长沙市望城经济技术开发区市场监督管理局",
        "高新开发区": "长沙市高新开发区市场监督管理局",
        "起步区": "济南市起步区市场监督管理局",
        "南昌经济技术开发区": "南昌市南昌经济技术开发区市场监督管理局",
        "走马岭街": "武汉市东西湖区市场监督管理局",
        "京东华中电商产业园": "武汉市东西湖区市场监督管理局",
        "京东亚洲一号": "西安市市场监督管理局浐灞国际港分局",
        "国际港务区": "西安市市场监督管理局浐灞国际港分局",
        "掌起镇": "慈溪市市场监督管理局",
        "塘厦": "东莞市市场监督管理局塘厦分局",
        "宿豫区": "宿迁市宿豫区市场监督管理局",
        "沭阳县": "宿迁市沭阳县市场监督管理局",
        "沪太路": "上海市静安区市场监督管理局",
        "罗城县": "河池市罗城仫佬族自治县市场监督管理局",
        "前湾新区": "宁波市市场监督管理局前湾新区分局",
        "义乌市": "义乌市市场监督管理局",
        "东明县": "菏泽市东明县市场监督管理局",
        "洛宁县": "洛阳市洛宁县市场监督管理局",
        "绿春县": "红河哈尼族彝族自治州绿春县市场监督管理局",
    }

    def lookup_by_address(self, address: str) -> str | None:
        if not address or address in ("暂无", "/", "None", ""):
            return None

        for pattern, bureau in self.SPECIAL_RULES.items():
            if pattern in address:
                return bureau

        result = self._generic_address_match(address)
        if result:
            return result

        logger.warning(f"Cannot determine bureau from address: {address[:50]}")
        return None

    def _generic_address_match(self, addr: str) -> str | None:
        m = re.match(r"((?:北京|上海|天津|重庆)市)(.{2,8}?[区县])", addr)
        if m:
            return f"{m.group(1)}{m.group(2)}市场监督管理局"

        m = re.match(r".*?省(.{2,6}?[市州盟])(.{2,8}?[区县旗])", addr)
        if m:
            return f"{m.group(1)}{m.group(2)}市场监督管理局"

        m = re.match(r".*?省(.{2,6}?[县旗])", addr)
        if m:
            return f"{m.group(1)}市场监督管理局"

        m = re.match(r".*?省(.{2,10}?自治[州县])(.{2,8}?[县旗])", addr)
        if m:
            return f"{m.group(1)}{m.group(2)}市场监督管理局"

        m = re.match(r"(.{2,6}?[市州盟])(.{2,8}?[区县旗开发区])", addr)
        if m:
            city, district = m.group(1), m.group(2)
            if "经济技术开发区" in district or "高新区" in district or "高新开发" in district:
                return f"{city}{district}市场监督管理局"
            return f"{city}{district}市场监督管理局"

        m = re.match(r"(.{2,6}?[市州盟])", addr)
        if m:
            return f"{m.group(1)}市场监督管理局"

        return None


def create_service(config) -> BureauDB:
    """Create a BureauDB instance (always local, no API)."""
    return BureauDB()
