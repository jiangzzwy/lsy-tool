"""Flask web UI for case transfer letter batch generation tool.

Run: python3 web.py
Opens http://localhost:5002 in browser.
"""

import json
import os
import platform
import shutil
import subprocess
import sys
import threading
import traceback
from collections import Counter
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file

import config
from api_client import BureauDB, MockApiClient
from excel_parser import parse_excel, split_rows
from word_generator import generate_word_docs
from ledger_generator import generate_ledger

app = Flask(__name__, template_folder="web/templates", static_folder="web/static")

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"

job_lock = threading.Lock()
job_state = {
    "running": False, "step": 1, "progress": 0.0, "progress_label": "",
    "status": "ready", "status_color": "green", "logs": [],
    "stats": {"db": 0, "need": "--", "ok": "--", "missing": "--",
              "word": "--", "ledger": "--", "err": "--", "cls": "--"},
    "parsed_items": None, "parsed_rows": None, "db_count": 0,
    "missing_info": None, "missing_count": 0,
}


def _reset_state():
    job_state["running"] = False
    job_state["step"] = 1
    job_state["progress"] = 0.0
    job_state["progress_label"] = ""
    job_state["status"] = "ready"
    job_state["status_color"] = "green"
    job_state["logs"] = []
    job_state["stats"] = {"db": 0, "need": "--", "ok": "--", "missing": "--",
                          "word": "--", "ledger": "--", "err": "--", "cls": "--"}
    job_state["parsed_items"] = None
    job_state["parsed_rows"] = None
    job_state["missing_info"] = None
    job_state["missing_count"] = 0


def _log(msg):
    job_state["logs"].append(msg)
    if len(job_state["logs"]) > 200:
        job_state["logs"] = job_state["logs"][-200:]


def _update_stats(**kw):
    for k, v in kw.items():
        job_state["stats"][k] = v


def _set_progress(pct, label=""):
    job_state["progress"] = pct
    job_state["progress_label"] = label


def _set_status(text, color="green"):
    job_state["status"] = text
    job_state["status_color"] = color


def _set_step(n):
    job_state["step"] = n


def _resolve_templates(src):
    sp = Path(src)
    cands = [sp.parent, BASE_DIR / "demands"]
    out = {}
    for cat, rel in config.TEMPLATES.items():
        name = Path(rel).name
        for c in cands:
            t = c / name
            if t.exists():
                out[cat] = str(t)
                break
    return out


def _find_default_excel():
    d = BASE_DIR / "demands"
    if d.exists():
        for f in sorted(d.iterdir(), key=lambda x: x.name):
            if f.suffix in (".xlsx", ".xls") and "工单" in f.name and not f.name.startswith("~$"):
                return str(f)
    return ""


def run_precheck(src, out):
    try:
        if not src or not Path(src).exists():
            _log("ERROR: source file not found")
            _set_status("file not found", "red")
            return
        if not out:
            out = str(OUTPUT_DIR)
        Path(out).mkdir(parents=True, exist_ok=True)

        _set_status("prechecking...", "yellow")
        _set_progress(0.15, "Parsing Excel...")
        _log("Parsing Excel...")
        rows = parse_excel(src)
        stats = Counter(r.classification for r in rows)
        for k, v in stats.items():
            _log(f"  {k}: {v}")

        _set_progress(0.35, "Splitting rows...")
        _log("Splitting rows...")
        items = split_rows(rows)
        _log(f"  Total: {len(items)}")

        job_state["parsed_items"] = items
        job_state["parsed_rows"] = rows

        _set_progress(0.55, "Checking database...")
        _log("Checking local database...")
        db = BureauDB()
        db_count = db.count()
        coverage = db.check_coverage(items)
        db.close()

        _log(f"  DB records: {db_count}")
        _log(f"  Need bureau: {coverage['total']}")
        _log(f"  Covered: {coverage['covered']} ({coverage['coverage_pct']:.0f}%)")
        _log(f"  Missing: {len(coverage['missing'])}")

        _update_stats(db=db_count, need=coverage["total"], ok=coverage["covered"], missing=len(coverage["missing"]))
        job_state["db_count"] = db_count

        if coverage["missing"]:
            job_state["missing_info"] = coverage["missing"]
            job_state["missing_count"] = len(coverage["missing"])
            _log(f"  >>> {len(coverage['missing'])} bureaus missing, need supplement")

            # Auto-export missing list as Excel
            tpl_path = Path(out) / "待补充登记机关.xlsx"
            db2 = BureauDB()
            db2.export_template(coverage["missing"], str(tpl_path))
            db2.close()
            _log(f"  >>> 已生成补充清单: {tpl_path}")
            _log(f"  >>> 请在清单中填写登记机关后，通过映射关系表的「导入」功能导入，再重新预检")

            _set_progress(1.0, "Missing " + str(job_state["missing_count"]))
            _set_status("needs_supplement", "yellow")
            _set_step(1)
        else:
            _log("All bureaus covered!")
            _set_progress(1.0, "Ready")
            _set_status("all covered", "green")
            _set_step(2)
    except Exception as e:
        _log(f"FATAL: {e}")
        _log(traceback.format_exc())
        _set_status("precheck failed", "red")
    finally:
        job_state["running"] = False


def run_import(import_path, out):
    try:
        if not import_path or not Path(import_path).exists():
            _log("ERROR: import file not found")
            _set_status("file not found", "red")
            return

        _set_status("importing...", "yellow")
        _set_progress(0.3, "Importing data...")

        db = BureauDB()
        result = db.import_full_replace(import_path)
        db.close()

        _log(f"Imported: {result['imported']} entries, {result['skipped']} skipped, {result['deleted']} old records replaced")

        db_count = BureauDB().count()
        _update_stats(db=db_count)
        _set_progress(0.6, "Re-checking...")

        if job_state["parsed_items"]:
            db2 = BureauDB()
            coverage = db2.check_coverage(job_state["parsed_items"])
            db2.close()

            _log(f"Re-check: {coverage['covered']}/{coverage['total']} covered")
            _log(f"Still missing: {len(coverage['missing'])}")

            _update_stats(ok=coverage["covered"], missing=len(coverage["missing"]))

            if coverage["missing"]:
                job_state["missing_info"] = coverage["missing"]
                job_state["missing_count"] = len(coverage["missing"])
                _log(f"Re-check: still {len(coverage['missing'])} missing")
                _log(f"  >>> 请继续补充登记机关后重新导入")

                _set_status("needs_supplement", "yellow")
                _set_step(1)
            else:
                _set_status("all covered", "green")
                _set_progress(1.0, "Ready")
                _set_step(2)
    except Exception as e:
        _log(f"FATAL: {e}")
        _log(traceback.format_exc())
        _set_status("import failed", "red")
    finally:
        job_state["running"] = False


def run_generation(out):
    try:
        if not out:
            out = str(OUTPUT_DIR)
        Path(out).mkdir(parents=True, exist_ok=True)

        items = job_state["parsed_items"]
        if not items:
            _log("ERROR: please precheck first")
            _set_status("please precheck", "red")
            return

        tpls = _resolve_templates(job_state.get("source_excel", ""))
        if not tpls:
            _log("ERROR: templates not found")
            return

        db = BureauDB()
        _set_status("generating...", "yellow")
        _set_progress(0.05, "Generating Word...")
        _log("Generating Word documents...")

        total = len(items)
        word_results = []
        for i, item in enumerate(items):
            cls = item.row.classification
            template_path = tpls.get(cls)
            if not template_path:
                continue
            try:
                from word_generator import _fill_template, _build_doc_filename
                doc = _fill_template(template_path, item, db)
                filename = _build_doc_filename(item)
                cls_dir = Path(out) / cls
                cls_dir.mkdir(parents=True, exist_ok=True)
                doc.save(str(cls_dir / f"{filename}.docx"))
                word_results.append({"filename": "ok"})
            except Exception as e:
                word_results.append({"error": str(e)})

            if (i + 1) % 20 == 0 or i == total - 1:
                pct = 0.05 + 0.85 * (i + 1) / total
                _set_progress(pct, f"Word {i+1}/{total}")

        ok = sum(1 for r in word_results if r.get("filename"))
        err = sum(1 for r in word_results if r.get("error"))
        _log(f"  Word: {ok} success, {err} failed")

        _set_progress(0.92, "Generating ledger...")
        _log("Generating ledger...")
        lp = generate_ledger(items, out, db, word_results=word_results)
        _log(f"  Ledger: {lp}")
        db.close()

        _set_progress(1.0, "Done")
        _set_status(f"done: {ok} Word + 1 ledger", "green")

        _update_stats(word=ok, ledger=1, err=err,
                       cls="  ".join(f"{k}:{v}" for k, v in Counter(
                           i.row.classification for i in items).items() if k != "other"))
    except Exception as e:
        _log(f"FATAL: {e}")
        _log(traceback.format_exc())
        _set_status("generation failed", "red")
    finally:
        job_state["running"] = False



class Api:
    def fileDialog(self):
        import webview
        result = webview.windows[0].create_file_dialog(
            webview.FileDialog.OPEN, file_types=('Excel Files (*.xlsx;*.xls)',)
        )
        if result and result[0]:
            return result[0]
        return ''

    def dirDialog(self):
        import webview
        result = webview.windows[0].create_file_dialog(webview.FileDialog.FOLDER)
        if result and result[0]:
            return result[0]
        return ''

    def openDir(self, path):
        import subprocess, platform
        s = platform.system()
        p = str(path) if path else ''
        if not p:
            return
        if s == 'Darwin':
            subprocess.Popen(['open', p])
        elif s == 'Windows':
            import os
            os.startfile(p)

    def downloadFile(self, url_path):
        """Open the output directory so user can find the exported file."""
        try:
            import subprocess, platform
            s = platform.system()
            if s == 'Darwin':
                subprocess.Popen(['open', str(OUTPUT_DIR)])
            elif s == 'Windows':
                import os
                os.startfile(str(OUTPUT_DIR))
            return True
        except Exception as e:
            print(f"downloadFile error: {e}")
        return False

    def confirmMissing(self, count):
        import webview
        msg = '缺失 ' + str(count) + ' 条企业登记机关数据，是否需要补充？'
        result = webview.windows[0].create_confirmation_dialog('缺失登记机关数据', msg)
        if result:
            _set_step(2)
            _set_status("ready for import", "yellow")
            return True
        else:
            _set_step(1)
            _set_status("cancelled supplement", "green")
            return False

api = Api()



@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/state")
def api_state():
    return jsonify({
        "running": job_state["running"], "step": job_state["step"],
        "progress": job_state["progress"], "progress_label": job_state["progress_label"],
        "status": job_state["status"], "status_color": job_state["status_color"],
        "stats": job_state["stats"], "logs": job_state["logs"][-100:],
        "missing_count": job_state["missing_count"],
        "missing_list": job_state.get("missing_info") or [],
    })

@app.route("/api/precheck", methods=["POST"])
def api_precheck():
    if job_state["running"]:
        return jsonify({"error": "task running"}), 400
    data = request.json or {}
    src = data.get("source_excel", _find_default_excel())
    out = data.get("output_dir", str(OUTPUT_DIR))
    job_state["running"] = True
    job_state["logs"] = []
    job_state["source_excel"] = src
    threading.Thread(target=run_precheck, args=(src, out), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/import", methods=["POST"])
def api_import():
    if job_state["running"]:
        return jsonify({"error": "task running"}), 400
    if job_state["step"] not in (1, 2):
        return jsonify({"error": "请先执行预检"}), 400
    if not job_state.get("missing_info"):
        return jsonify({"error": "无需补充数据"}), 400
    data = request.json or {}
    import_path = data.get("import_path", "")
    if not import_path or not Path(import_path).exists():
        return jsonify({"error": "请选择有效的补充模板文件"}), 400
    out = data.get("output_dir", str(OUTPUT_DIR))
    job_state["running"] = True
    threading.Thread(target=run_import, args=(import_path, out), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/generate", methods=["POST"])
def api_generate():
    if job_state["running"]:
        return jsonify({"error": "task running"}), 400
    if job_state["step"] < 2:
        return jsonify({"error": "请先完成预检"}), 400
    if not job_state.get("parsed_items"):
        return jsonify({"error": "请先执行预检"}), 400
    data = request.json or {}
    out = data.get("output_dir", str(OUTPUT_DIR))
    job_state["running"] = True
    threading.Thread(target=run_generation, args=(out,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/reset", methods=["POST"])
def api_reset():
    _reset_state()
    return jsonify({"ok": True})

@app.route("/api/clear_log", methods=["POST"])
def api_clear_log():
    job_state["logs"] = []
    return jsonify({"ok": True})

@app.route("/api/export_template", methods=["POST"])
def api_export_template():
    data = request.json or {}
    save_dir = data.get("save_dir", str(OUTPUT_DIR))
    if not job_state["missing_info"]:
        return jsonify({"error": "no missing data"}), 400
    Path(save_dir).mkdir(parents=True, exist_ok=True)
    tpl_path = Path(save_dir) / "missing_bureau_template.xlsx"
    db = BureauDB()
    db.export_template(job_state["missing_info"], str(tpl_path))
    db.close()
    _log(f"Template exported: {tpl_path}")
    return jsonify({"ok": True, "path": str(tpl_path), "download": "/api/download/missing_bureau_template.xlsx"})

@app.route("/api/confirm_supplement", methods=["POST"])
def api_confirm_supplement():
    _set_step(1)
    _set_status("ready for import", "yellow")
    return jsonify({"ok": True})

@app.route("/api/cancel_supplement", methods=["POST"])
def api_cancel_supplement():
    _set_step(1)
    _set_status("cancelled", "green")
    return jsonify({"ok": True})

@app.route("/api/export_db", methods=["POST"])
def api_export_db():
    db = BureauDB()
    rows = db._db.execute("SELECT credit_code, bureau, company, source, updated_at FROM bureau_cache ORDER BY credit_code").fetchall()
    db.close()
    if not rows:
        return jsonify({"error": "no records"}), 400
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业登记机关"
    headers = ["统一社会信用代码", "登记机关", "企业名称", "来源", "更新时间"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    for row_idx, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row_idx, col, val)
    dest = OUTPUT_DIR / "bureau_database.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))
    wb.close()
    return jsonify({"ok": True, "count": len(rows), "download": f"/api/download/bureau_database.xlsx"})


@app.route("/api/mapping_records")
def api_mapping_records():
    db = BureauDB()
    records = db.get_all_records()
    db.close()
    return jsonify({"records": records, "count": len(records)})


@app.route("/api/mapping/update", methods=["POST"])
def api_mapping_update():
    """Add or update a single mapping record."""
    data = request.json or {}
    cc = (data.get("credit_code") or "").strip()
    bureau = (data.get("bureau") or "").strip()
    company = (data.get("company") or "").strip()
    if not cc or not bureau:
        return jsonify({"error": "信用代码和登记机关不能为空"}), 400
    db = BureauDB()
    try:
        db._db.execute(
            """INSERT OR REPLACE INTO bureau_cache (credit_code, bureau, company, source, updated_at)
               VALUES (?, ?, ?, 'manual', datetime('now', 'localtime'))""",
            (cc, bureau, company),
        )
        db._db.commit()
        db.close()
        return jsonify({"ok": True})
    except Exception as e:
        db.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/mapping/delete", methods=["POST"])
def api_mapping_delete():
    """Delete a single mapping record by credit_code."""
    data = request.json or {}
    cc = (data.get("credit_code") or "").strip()
    if not cc:
        return jsonify({"error": "信用代码不能为空"}), 400
    db = BureauDB()
    try:
        db._db.execute("DELETE FROM bureau_cache WHERE credit_code = ?", (cc,))
        db._db.commit()
        db.close()
        return jsonify({"ok": True})
    except Exception as e:
        db.close()
        return jsonify({"error": str(e)}), 500


@app.route("/api/import_update", methods=["POST"])
def api_import_update():
    """Import and merge: update existing records, add new ones.
    
    Used by the '导入更新' button in mapping table toolbar.
    The import file format matches export_db output:
    Column A=credit_code, B=bureau, C=company (or A=credit_code, B=company, D=bureau)
    """
    data = request.json or {}
    import_path = data.get("import_path", "")
    if not import_path or not Path(import_path).exists():
        return jsonify({"error": "请选择有效的导入文件"}), 400
    
    db = BureauDB()
    try:
        result = db.import_template(import_path)
        db.close()
        return jsonify({
            "ok": True,
            "imported": result["imported"],
            "skipped": result["skipped"],
            "errors": result.get("errors", [])
        })
    except Exception as e:
        db.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/download/<filename>")
def api_download(filename):
    fpath = OUTPUT_DIR / filename
    if fpath.exists():
        return send_file(str(fpath), as_attachment=True)
    return jsonify({"error": "file not found"}), 404

@app.route("/api/default_excel")
def api_default_excel():
    return jsonify({"path": _find_default_excel()})


if __name__ == "__main__":
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    import threading, time
    server_thread = threading.Thread(
        target=lambda: app.run(host="127.0.0.1", port=5004, debug=False, use_reloader=False),
        daemon=True
    )
    server_thread.start()
    for _ in range(20):
        try:
            import urllib.request
            urllib.request.urlopen("http://127.0.0.1:5004/")
            break
        except Exception:
            time.sleep(0.5)
    import webview
    window = webview.create_window(
        title="案件线索移送函批量生成",
        url="http://127.0.0.1:5004/",
        width=960, height=720, resizable=True, min_size=(720, 560),
        background_color='#1a1b23',
        js_api=api,
    )
    webview.start()
